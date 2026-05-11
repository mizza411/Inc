"""Decode Strategy 15 payload + dump Telecoms xlsx to UTF-8 text (run from this folder)."""
import pathlib

try:
    import openpyxl
except ImportError:
    raise SystemExit("Install openpyxl: py -m pip install openpyxl")

root = pathlib.Path(__file__).resolve().parent
lines: list[str] = []

p = root / "strategy15_prompt_1a_payload.txt"
raw = p.read_bytes()
for enc in ("utf-8-sig", "utf-16-le", "utf-16", "cp1252", "latin-1"):
    try:
        t = raw.decode(enc)
        lines.append(f"=== strategy15_prompt_1a_payload.txt (encoding {enc}) ===")
        lines.append(t[:120000])
        lines.append("")
        break
    except UnicodeDecodeError:
        continue
else:
    lines.append(f"(Could not decode payload as text; {len(raw)} bytes)")

xlsx = root / "Nu" / "Telecoms_Q3_2025.xlsx"
lines.append(f"=== XLSX: {xlsx.name} ===")
wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
for sn in wb.sheetnames:
    ws = wb[sn]
    lines.append(f"\n--- SHEET: {sn} (max_row={ws.max_row}) ---")
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= 200:
            lines.append("... (truncated after 200 rows)")
            break
        lines.append(str(row))
wb.close()

out = root / "_telecoms_q3_dump.txt"
out.write_text("\n".join(lines), encoding="utf-8")
print("Wrote:", out)
